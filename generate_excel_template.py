# -*- coding: utf-8 -*-
"""
generate_excel_template.py
---------------------------
비개발자(데이터 입력 담당자)가 엑셀로 병원 데이터를 채울 수 있도록
빈 템플릿(예제 1행 포함)을 생성하는 스크립트.

실행:
    python generate_excel_template.py

생성 결과:
    병원데이터_템플릿.xlsx  (같은 폴더에 생성)

이 템플릿을 채운 뒤 excel_to_db.py로 실제 DB(SQLite)에 적재합니다.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from database import EQUIPMENT_TYPES, CHECKUP_TYPE_NAMES, EXAM_ITEM_MASTER
from korea_regions import REGION_MAP, SIDO_LIST

OUTPUT_PATH = "병원데이터_템플릿.xlsx"

FONT_NAME = "Arial"
HEADER_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 예제 입력행 표시
HEADER_FONT = Font(name=FONT_NAME, bold=True)
NORMAL_FONT = Font(name=FONT_NAME)


def _style_header(ws, headers):
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(title) * 2)


def _style_example_row(ws, row_idx, values):
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.fill = INPUT_FILL


def build_template():
    wb = Workbook()

    # -----------------------------------------------------------------
    # 0. 작성 안내
    # -----------------------------------------------------------------
    guide = wb.active
    guide.title = "0_작성안내"
    guide.column_dimensions["A"].width = 100
    lines = [
        "■ 이 템플릿은 총 4개의 데이터 입력 시트로 구성되어 있습니다.",
        "  1_병원마스터 : 병원 기본정보 (필수 - 다른 모든 시트가 여기의 '병원ID'를 참조합니다)",
        "  2_보유장비   : 병원별 보유 장비 (CT/MRI/PET-CT)",
        "  3_검진유형가격 : 병원별 종합검진/일반검진 가격 및 소요시간",
        "  4_검사항목   : 병원별로 제공 가능한 세부 검사항목",
        "",
        "■ 작성 규칙",
        "  - 노란색으로 칠해진 첫 데이터행은 '예시'입니다. 실제 데이터를 그 아래부터 이어서 입력하세요.",
        "  - '병원ID'는 병원마다 고유한 숫자여야 하며, 4개 시트 모두 동일한 병원ID로 연결됩니다.",
        "    (예: 병원ID 1번 병원의 장비/가격/검사항목은 각 시트에서 모두 병원ID=1로 입력)",
        "  - 시/도, 검진유형, 장비종류, 검사항목명은 드롭다운으로 제공되는 값 중에서만 선택하세요.",
        "    (직접 다른 문자열을 입력하면 변환 시 오류로 처리되어 누락될 수 있습니다)",
        "  - 시/군/구는 '참고_시도구군' 시트에서 정확한 표기를 복사해서 사용하세요.",
        "  - 작성이 끝나면 파일명을 유지한 채 저장하고, excel_to_db.py 스크립트로 변환하세요.",
        "",
        "■ 문의: 이 템플릿 구조를 변경(열 추가/삭제)하면 변환 스크립트도 함께 수정해야 합니다.",
    ]
    for i, line in enumerate(lines, start=1):
        c = guide.cell(row=i, column=1, value=line)
        c.font = Font(name=FONT_NAME, bold=line.startswith("■"))

    # -----------------------------------------------------------------
    # 1. 병원마스터
    # -----------------------------------------------------------------
    ws1 = wb.create_sheet("1_병원마스터")
    headers1 = [
        "병원ID", "병원명", "시도", "시군구", "주소", "전화번호",
        "홈페이지", "설립연도", "인증(콤마구분)", "소개",
    ]
    _style_header(ws1, headers1)
    _style_example_row(ws1, 2, [
        1, "샘플종합검진센터", "서울", "강남구", "서울 강남구 테헤란로 1",
        "02-0000-0000", "https://example.com", 2015, "국가건강검진기관,JCI인증",
        "샘플종합검진센터는 서울 강남구의 건강검진 전문기관입니다.",
    ])

    dv_sido = DataValidation(type="list", formula1=f'"{",".join(SIDO_LIST)}"', allow_blank=True)
    ws1.add_data_validation(dv_sido)
    dv_sido.add(f"C2:C1000")

    # -----------------------------------------------------------------
    # 2. 보유장비
    # -----------------------------------------------------------------
    ws2 = wb.create_sheet("2_보유장비")
    headers2 = ["병원ID", "장비종류", "스펙(선택)", "도입연도(선택)"]
    _style_header(ws2, headers2)
    _style_example_row(ws2, 2, [1, "CT", "128채널 CT", 2020])

    dv_equipment = DataValidation(
        type="list", formula1=f'"{",".join(EQUIPMENT_TYPES)}"', allow_blank=True
    )
    ws2.add_data_validation(dv_equipment)
    dv_equipment.add("B2:B1000")

    # -----------------------------------------------------------------
    # 3. 검진유형가격
    # -----------------------------------------------------------------
    ws3 = wb.create_sheet("3_검진유형가격")
    headers3 = ["병원ID", "검진유형", "최소가격(원)", "최대가격(원)", "평균소요시간(시간)"]
    _style_header(ws3, headers3)
    _style_example_row(ws3, 2, [1, "종합검진", 400000, 900000, 3.5])

    dv_checkup = DataValidation(
        type="list", formula1=f'"{",".join(CHECKUP_TYPE_NAMES)}"', allow_blank=True
    )
    ws3.add_data_validation(dv_checkup)
    dv_checkup.add("B2:B1000")

    # -----------------------------------------------------------------
    # 4. 검사항목
    # -----------------------------------------------------------------
    ws4 = wb.create_sheet("4_검사항목")
    headers4 = ["병원ID", "검사항목명"]
    _style_header(ws4, headers4)
    _style_example_row(ws4, 2, [1, "위내시경"])

    # 검사항목명은 목록이 길어서(22개) 인라인 리스트 대신 참고 시트 범위를 참조
    ref_items = wb.create_sheet("참고_검사항목목록")
    ref_items.sheet_state = "hidden"
    ref_items.cell(row=1, column=1, value="검사항목명").font = HEADER_FONT
    for i, (_category, item_name) in enumerate(EXAM_ITEM_MASTER, start=2):
        ref_items.cell(row=i, column=1, value=item_name).font = NORMAL_FONT

    dv_items = DataValidation(
        type="list",
        formula1=f"'참고_검사항목목록'!$A$2:$A${len(EXAM_ITEM_MASTER) + 1}",
        allow_blank=True,
    )
    ws4.add_data_validation(dv_items)
    dv_items.add("B2:B1000")

    # -----------------------------------------------------------------
    # 참고_시도구군 (읽기전용 참고자료, 시군구 정확한 표기 복사용)
    # -----------------------------------------------------------------
    ref_region = wb.create_sheet("참고_시도구군")
    ref_region.cell(row=1, column=1, value="시도").font = HEADER_FONT
    ref_region.cell(row=1, column=2, value="시군구").font = HEADER_FONT
    row = 2
    for si, gu_list in REGION_MAP.items():
        for gu in gu_list:
            ref_region.cell(row=row, column=1, value=si).font = NORMAL_FONT
            ref_region.cell(row=row, column=2, value=gu).font = NORMAL_FONT
            row += 1
    ref_region.column_dimensions["A"].width = 12
    ref_region.column_dimensions["B"].width = 20

    wb.save(OUTPUT_PATH)
    print(f"템플릿 생성 완료: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_template()
